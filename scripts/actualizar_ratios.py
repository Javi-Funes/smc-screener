"""
ACTUALIZAR RATIOS CEDEARs
=========================
Fuente 1: Comafi       -> Acciones USA + globales (~300 tickers)
Fuente 2: Caja de Valores -> ETFs + Acciones Brasil (~55 tickers)

Los links al Excel se obtienen dinamicamente scrapeando cada pagina,
asi si cambia el nombre del archivo lo encontramos igual.

Corre: Domingos 8AM ARG (o manualmente)
Output: results/ratios_cedears.json + results/ratios_cedears.csv

Formula precio CEDEAR en pesos:
  precio_ars = (precio_usd / ratio) * ccl
  Ejemplo BABA: ($136.29 / 9) * $1,455 = $22,044 por CEDEAR
"""

import requests
import pandas as pd
import json
import os
import re
from datetime import datetime
from io import BytesIO
from openpyxl import load_workbook

HEADERS = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'}

OUTPUT_JSON = 'results/ratios_cedears.json'
OUTPUT_CSV  = 'results/ratios_cedears.csv'

URL_COMAFI_PAGE = 'https://www.comafi.com.ar/custodiaglobal/Programas-CEDEARs-2483.note.aspx'
URL_COMAFI_BASE = 'https://www.comafi.com.ar'
URL_CAJVAL_PAGE = 'https://cajadevalores.com.ar/Servicios/Cedears'
URL_CAJVAL_BASE = 'https://cajadevalores.com.ar'

def get_excel_url(page_url, base_url):
    try:
        r = requests.get(page_url, headers=HEADERS, timeout=15)
        if r.status_code != 200:
            return None
        matches = re.findall(r'href=["\']([^"\']*\.xlsx[^"\']*)["\']', r.text, re.IGNORECASE)
        if not matches:
            return None
        link = matches[0]
        if link.startswith('http'):
            return link
        elif link.startswith('/'):
            return base_url + link
        else:
            return base_url + '/' + link
    except Exception as e:
        print(f"  Error scrapeando {page_url}: {e}")
        return None

def descargar(url, nombre):
    try:
        print(f"  URL: {url}")
        r = requests.get(url, timeout=20, headers=HEADERS)
        if r.status_code == 200 and len(r.content) > 1000:
            print(f"  OK: {len(r.content):,} bytes")
            return r.content
        print(f"  Error HTTP {r.status_code}")
        return None
    except Exception as e:
        print(f"  Error: {e}")
        return None

def parse_ratio(raw):
    raw = str(raw).strip()
    if not raw or raw.upper() in ('NAN', 'N/A', '-', '', 'NONE'):
        return None
    try:
        m = re.match(r'(\d+\.?\d*)\s*[:/]\s*(\d+\.?\d*)', raw)
        if m:
            num, den = float(m.group(1)), float(m.group(2))
            return round(num / den, 4) if den != 0 else None
        val = float(re.sub(r'[^\d.]', '', raw))
        return val if val > 0 else None
    except Exception:
        return None

def parse_comafi(content):
    """
    Estructura real del Excel Comafi:
      Hoja: 'LISTA TOTAL DE CEDEARS'
      Header fila 8:
        Col B: DENOMINACION DEL PROGRAMA
        Col C: Identificacion Mercado  (ticker)
        Col H: Ratio Cedear/Accion
    """
    ratios = {}
    try:
        wb = load_workbook(BytesIO(content), read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        print(f"  Comafi hoja: '{wb.sheetnames[0]}'")

        COL_NOMBRE = 2  # B
        COL_TICKER = 3  # C
        COL_RATIO  = 8  # H

        for row in ws.iter_rows(min_row=9, values_only=True):  # datos desde fila 9
            if not row or len(row) < COL_RATIO:
                continue
            ticker    = str(row[COL_TICKER-1]).strip().upper() if row[COL_TICKER-1] else ''
            nombre    = str(row[COL_NOMBRE-1]).strip()         if row[COL_NOMBRE-1] else ''
            ratio_raw = str(row[COL_RATIO -1]).strip()         if row[COL_RATIO -1] else ''

            if not ticker or ticker in ('NAN', 'IDENTIFICACIÓN MERCADO', ''):
                continue

            ratio = parse_ratio(ratio_raw)
            if ratio is None:
                continue

            ratios[ticker] = {
                'ratio':  ratio,
                'nombre': nombre,
                'byma':   ticker,
                'tipo':   'Accion',
                'fuente': 'Comafi',
            }

        print(f"  Comafi: {len(ratios)} tickers parseados")
    except Exception as e:
        print(f"  Comafi ERROR: {e}")
    return ratios

def parse_cajavaloroes(content):
    """
    Estructura real: hoja 'Actualizaciones 12-2'
    Header en fila 1, columnas detectadas por nombre.
    """
    ETF_LIST = {'SPY','QQQ','IWM','EEM','XLF','XLE','DIA','EWZ','ARKK',
                'SH','GLD','ETHA','URA','SMH','SPXL','XLU','CIBR','TQQQ','VXX','ITA'}
    ratios = {}
    try:
        wb = load_workbook(BytesIO(content), read_only=True, data_only=True)
        print(f"  CajaValores hojas: {wb.sheetnames}")

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue

            # Buscar fila header
            header_idx = None
            header = None
            for i, row in enumerate(rows[:10]):
                if any('ratio' in str(c).lower() for c in row if c):
                    header_idx = i
                    header = [str(c).strip() if c else '' for c in row]
                    break
            if header_idx is None:
                continue

            print(f"  CajaValores hoja '{sheet_name}': header fila {header_idx+1}")

            def find_col(kws):
                for j, h in enumerate(header):
                    if all(k.lower() in h.lower() for k in kws):
                        return j
                return None

            col_ticker = find_col(['ticker','origen']) or find_col(['ticker'])
            col_byma   = find_col(['byma']) or find_col(['símbolo']) or find_col(['simbolo'])
            col_ratio  = find_col(['ratio'])
            col_nombre = find_col(['cedear','etf']) or find_col(['cedear','accion']) or find_col(['nombre'])

            if col_ratio is None:
                continue

            for row in rows[header_idx+1:]:
                if not any(row):
                    continue
                ticker_origen = str(row[col_ticker]).strip().upper() if col_ticker is not None and row[col_ticker] else ''
                ticker_byma   = str(row[col_byma  ]).strip().upper() if col_byma   is not None and row[col_byma  ] else ''
                ratio_raw     = str(row[col_ratio ]).strip()          if col_ratio  is not None and row[col_ratio ] else ''
                nombre        = str(row[col_nombre]).strip()          if col_nombre is not None and row[col_nombre] else ''

                ticker = ticker_origen or ticker_byma
                if not ticker or ticker.upper() in ('NAN',''):
                    continue

                ratio = parse_ratio(ratio_raw)
                if ratio is None:
                    continue

                if any(ticker.endswith(s) for s in ['3','4','11']):
                    tipo = 'Brasil'
                elif ticker in ETF_LIST:
                    tipo = 'ETF'
                else:
                    tipo = 'Accion'

                ratios[ticker] = {
                    'ratio':  ratio,
                    'nombre': nombre,
                    'byma':   ticker_byma or ticker,
                    'tipo':   tipo,
                    'fuente': 'CajaValores',
                }

        print(f"  CajaValores: {len(ratios)} tickers parseados")
    except Exception as e:
        print(f"  CajaValores ERROR: {e}")
    return ratios

if __name__ == '__main__':
    inicio = datetime.now()
    print(f'\n{"="*60}')
    print(f'  ACTUALIZADOR DE RATIOS CEDEARs')
    print(f'  {inicio.strftime("%d/%m/%Y %H:%M")}')
    print(f'{"="*60}\n')

    os.makedirs('results', exist_ok=True)

    # 1. Obtener URLs dinamicamente
    print('[ 1/4 ] Buscando link Excel en pagina Comafi...')
    url_comafi = get_excel_url(URL_COMAFI_PAGE, URL_COMAFI_BASE)
    if url_comafi:
        print(f'  Link encontrado: {url_comafi}')
    else:
        url_comafi = URL_COMAFI_BASE + '/Multimedios/otros/7279.xlsx'
        print(f'  Usando fallback: {url_comafi}')

    print('\n[ 2/4 ] Buscando link Excel en pagina Caja de Valores...')
    url_cajval = get_excel_url(URL_CAJVAL_PAGE, URL_CAJVAL_BASE)
    if url_cajval:
        print(f'  Link encontrado: {url_cajval}')
    else:
        print('  No se encontro link en la pagina')

    # 2. Descargar
    print('\n[ 3/4 ] Descargando...')
    content_comafi = descargar(url_comafi, 'Comafi')       if url_comafi else None
    content_cajval = descargar(url_cajval, 'CajaValores')  if url_cajval else None

    # 3. Parsear
    print('\n[ 4/4 ] Parseando...')
    ratios_comafi = parse_comafi(content_comafi)       if content_comafi else {}
    ratios_cajval = parse_cajavaloroes(content_cajval) if content_cajval else {}

    # 4. Merge — CajaValores base, Comafi sobreescribe acciones
    ratios_final = {}
    for t, d in ratios_cajval.items():
        ratios_final[t] = d
    for t, d in ratios_comafi.items():
        ratios_final[t] = d  # Comafi gana en acciones

    # 5. Stats
    tipos = {}
    for d in ratios_final.values():
        t = d.get('tipo','?')
        tipos[t] = tipos.get(t,0) + 1

    print(f'\n  RESULTADO FINAL:')
    print(f'  Total tickers:    {len(ratios_final)}')
    for tipo, count in sorted(tipos.items()):
        print(f'  {tipo:<14}  {count}')
    print(f'  De Comafi:        {len(ratios_comafi)}')
    print(f'  De CajaValores:   {len(ratios_cajval)}')

    # 6. Guardar JSON
    output = {'_meta': {'actualizado': inicio.strftime('%Y-%m-%d %H:%M'),
                        'total': len(ratios_final), 'comafi': len(ratios_comafi),
                        'cajavaloroes': len(ratios_cajval)}, **ratios_final}
    with open(OUTPUT_JSON, 'w', encoding='utf-8') as f:
        json.dump(output, f, ensure_ascii=False, indent=2)
    print(f'\n  Guardado: {OUTPUT_JSON}')

    # 7. Guardar CSV
    rows = [{'ticker':t,'byma':d.get('byma',t),'ratio':d['ratio'],
             'nombre':d.get('nombre',''),'tipo':d['tipo'],'fuente':d['fuente']}
            for t,d in ratios_final.items()]
    pd.DataFrame(rows).sort_values(['tipo','ticker']).to_csv(OUTPUT_CSV, index=False)
    print(f'  Guardado: {OUTPUT_CSV}')

    # 8. Verificacion tickers clave
    print(f'\n  VERIFICACION TICKERS CLAVE:')
    for t in ['BABA','META','AAPL','MSFT','NVDA','AMZN','GOOGL','TSLA',
              'SPY','QQQ','VALE3','GGAL','YPF']:
        d = ratios_final.get(t)
        if d:
            print(f'  {t:<8}  ratio: {d["ratio"]:>6.1f}  tipo: {d["tipo"]:<8}  ({d["fuente"]})')
        else:
            print(f'  {t:<8}  NO ENCONTRADO')

    elapsed = (datetime.now() - inicio).seconds
    print(f'\n  Tiempo: {elapsed}s')
    print(f'{"="*60}')
